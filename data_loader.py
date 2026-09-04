import os
import io
import re
import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional
import openpyxl

DEFAULT_SEARCH_PATHS = [
    Path(r"C:\Users\alexc\Downloads\Alejandra\Alejandra"),
    Path(__file__).resolve().parent,
    Path(__file__).resolve().parent.parent,
    Path.cwd(),
]

def find_data_directory(custom_path: Optional[str] = None) -> Path:
    """Busca el directorio que contiene los archivos aplicacion.xlsm, Base.xlsx y Picto."""
    if custom_path and os.path.exists(custom_path):
        p = Path(custom_path)
        if (p / "aplicacion.xlsm").exists() or (p / "Base.xlsx").exists():
            return p

    for path in DEFAULT_SEARCH_PATHS:
        if path.exists():
            if (path / "aplicacion.xlsm").exists() and (path / "Base.xlsx").exists():
                return path

    for path in DEFAULT_SEARCH_PATHS:
        if path.exists() and (path / "aplicacion.xlsm").exists():
            return path

    return Path(r"C:\Users\alexc\Downloads\Alejandra\Alejandra")


def read_file_bytes_non_blocking(filepath: Path) -> io.BytesIO:
    """Lee el archivo en un búfer en memoria permitiendo lectura compartida si está abierto en Excel."""
    if not filepath.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {filepath}")
    
    with open(filepath, "rb") as f:
        data = f.read()
    return io.BytesIO(data)


def normalize_text(text: Any) -> str:
    """Normaliza texto para búsquedas y cruces insensibles a mayúsculas/espacios."""
    if text is None:
        return ""
    text_str = str(text).strip().upper()
    text_str = re.sub(r"\s+", " ", text_str)
    return text_str


def get_pictogram_catalog(picto_dir: Path) -> Dict[str, str]:
    """Escanea la carpeta Picto y crea un mapa de búsqueda por código GHS y nombre."""
    catalog = {}
    if not picto_dir.exists() or not picto_dir.is_dir():
        return catalog

    for file in picto_dir.glob("*.jpg"):
        fname = file.name
        norm = normalize_text(fname)
        catalog[norm] = fname
        match = re.search(r"(GHS\d{2})", norm)
        if match:
            ghs_code = match.group(1)
            catalog[ghs_code] = fname
            catalog[f"{ghs_code}.JPG"] = fname

    return catalog


def resolve_pictogram(pic_val: Any, picto_catalog: Dict[str, str], picto_dir: Path) -> Dict[str, Any]:
    """Resuelve la URL y metadatos de un pictograma para la etiqueta SGA."""
    if not pic_val:
        return {"has_image": False, "url": None, "label": "SIN IMAGEN", "code": None}

    val_str = str(pic_val).strip()
    norm = normalize_text(val_str)

    # Detectar si contiene 'SIN FOTO', 'SIN F', 'NO TIENE', 'N/A', etc.
    if any(k in norm for k in ["SIN FOTO", "SIN F", "NO TIENE", "NO APLICA", "N/A", "NINGUNO"]):
        return {"has_image": False, "url": None, "label": "SIN IMAGEN", "code": None}

    # Extraer código GHS (ej: GHS01, GHS06, etc.)
    match = re.search(r"(GHS\d{2})", norm)
    ghs_code = match.group(1) if match else None

    # Buscar archivo físico
    target_filename = None
    if ghs_code and (picto_dir / f"{ghs_code}.jpg").exists():
        target_filename = f"{ghs_code}.jpg"
    elif norm in picto_catalog and (picto_dir / picto_catalog[norm]).exists():
        target_filename = picto_catalog[norm]
    elif ghs_code and ghs_code in picto_catalog and (picto_dir / picto_catalog[ghs_code]).exists():
        target_filename = picto_catalog[ghs_code]
    else:
        # Búsqueda parcial
        for cat_norm, fname in picto_catalog.items():
            if (cat_norm in norm or norm in cat_norm) and (picto_dir / fname).exists():
                target_filename = fname
                break

    if target_filename:
        return {
            "has_image": True,
            "url": f"/picto/{target_filename}",
            "filename": target_filename,
            "label": val_str,
            "code": ghs_code or val_str
        }

    return {"has_image": False, "url": None, "label": "SIN IMAGEN", "code": None}


def load_base_catalog(base_path: Path, picto_dir: Path) -> Dict[str, Dict[str, Any]]:
    """Carga el maestro de productos desde Base.xlsx (hoja SGA)."""
    buf = read_file_bytes_non_blocking(base_path)
    wb = openpyxl.load_workbook(buf, data_only=True, read_only=True)
    
    sheet_name = "SGA" if "SGA" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    picto_catalog = get_pictogram_catalog(picto_dir)
    products = {}

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return products

    # Buscar fila de encabezados
    header_idx = 0
    header_map = {}
    for idx, row in enumerate(rows[:10]):
        row_str = [normalize_text(c) for c in row if c is not None]
        if "CODIGO" in row_str or "NOMBRE DEL PRODUCTO" in row_str:
            header_idx = idx
            for col_idx, col_name in enumerate(row):
                if col_name is not None:
                    header_map[normalize_text(col_name)] = col_idx
            break

    def get_val(row_tuple, col_name, default=""):
        idx = header_map.get(normalize_text(col_name))
        if idx is not None and idx < len(row_tuple) and row_tuple[idx] is not None:
            return row_tuple[idx]
        return default

    for row in rows[header_idx + 1:]:
        if not any(row):
            continue
        
        prod_name = get_val(row, "Nombre del Producto")
        if not prod_name or str(prod_name).strip() == "":
            continue

        codigo = str(get_val(row, "Codigo", "")).strip()
        um = str(get_val(row, "U/m", "KILO")).strip().upper()
        if not um:
            um = "KILO"
            
        frase_h = str(get_val(row, "FRASES H", "")).strip()
        frase_p = str(get_val(row, "FRASES P", "")).strip()
        palabra_adv = str(get_val(row, "PALABRA DE ADVERTENCIA", "")).strip().upper()
        
        pig1_val = get_val(row, "PIG1", "SIN FOTO")
        pig2_val = get_val(row, "PIG2", "SIN FOTO")
        pig3_val = get_val(row, "PIG3", "SIN FOTO")
        pig4_val = get_val(row, "PIG4", "SIN FOTO")

        pictograms = [
            resolve_pictogram(pig1_val, picto_catalog, picto_dir),
            resolve_pictogram(pig2_val, picto_catalog, picto_dir),
            resolve_pictogram(pig3_val, picto_catalog, picto_dir),
            resolve_pictogram(pig4_val, picto_catalog, picto_dir),
        ]

        prod_info = {
            "codigo": codigo,
            "nombre": str(prod_name).strip(),
            "um": um,
            "frase_h": frase_h,
            "frase_p": frase_p,
            "palabra_advertencia": palabra_adv,
            "pictogramas": pictograms,
            "tiene_sga": str(get_val(row, "Tiene SGA", "")).strip(),
            "tiene_hds": str(get_val(row, "TIENE HOJA DE SEGURIDAD", "")).strip(),
        }

        norm_key = normalize_text(prod_name)
        products[norm_key] = prod_info
        if codigo:
            products[f"COD:{codigo.upper()}"] = prod_info

    return products


def format_date_str(val: Any) -> Dict[str, str]:
    """Formatea fecha a formato ISO y formato latino legible."""
    if isinstance(val, (datetime.datetime, datetime.date)):
        iso_val = val.strftime("%Y-%m-%d")
        display_val = val.strftime("%d/%m/%Y")
        return {"iso": iso_val, "display": display_val}
    elif isinstance(val, str):
        val_str = val.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%d-%m-%Y"):
            try:
                dt = datetime.datetime.strptime(val_str, fmt)
                return {"iso": dt.strftime("%Y-%m-%d"), "display": dt.strftime("%d/%m/%Y")}
            except ValueError:
                pass
        return {"iso": val_str, "display": val_str}
    return {"iso": "", "display": "N/A"}


def load_applications(
    app_path: Path, 
    base_catalog: Dict[str, Dict[str, Any]]
) -> List[Dict[str, Any]]:
    """Carga y procesa la programación semanal desde aplicacion.xlsm (hoja Data)."""
    buf = read_file_bytes_non_blocking(app_path)
    wb = openpyxl.load_workbook(buf, data_only=True, read_only=True)
    
    sheet_name = "Data" if "Data" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx = -1
    header_map = {}
    for idx, row in enumerate(rows[:15]):
        row_str = [normalize_text(c) for c in row if c is not None]
        if "PRODUCTO" in row_str or "FECHA" in row_str or "LITROS AGUA" in row_str:
            header_idx = idx
            for col_idx, col_name in enumerate(row):
                if col_name is not None:
                    cleaned = normalize_text(col_name)
                    header_map[cleaned] = col_idx
            break

    if header_idx == -1:
        header_idx = 0
        for col_idx, col_name in enumerate(rows[0]):
            if col_name:
                header_map[normalize_text(col_name)] = col_idx

    def get_val(row_tuple, *col_aliases, default=None):
        for alias in col_aliases:
            idx = header_map.get(normalize_text(alias))
            if idx is not None and idx < len(row_tuple) and row_tuple[idx] is not None:
                return row_tuple[idx]
            norm_alias = normalize_text(alias)
            for h_key, h_idx in header_map.items():
                if norm_alias in h_key and h_idx < len(row_tuple) and row_tuple[h_idx] is not None:
                    return row_tuple[h_idx]
        return default

    applications = []
    item_counter = 0

    for row_num, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if not any(row):
            continue

        raw_prod = get_val(row, "PRODUCTO", "Producto", default="")
        if not raw_prod or str(raw_prod).strip() == "" or normalize_text(raw_prod) == "PRODUCTO":
            continue

        prod_name = str(raw_prod).strip()
        date_raw = get_val(row, "Fecha", "FECHA")
        date_info = format_date_str(date_raw)

        cultivo = str(get_val(row, "Cultivo", default="") or "").strip()
        bloque = str(get_val(row, "Bloque", default="") or "").strip()
        camas = get_val(row, "Camas", default="")
        reentrada = get_val(row, "Reentrada", default=0)
        categoria = get_val(row, "Categoria", "Categoría", default="")
        observaciones = str(get_val(row, "Observaciones", default="") or "").strip()
        semana = get_val(row, "Semana", default="")

        litros_raw = get_val(row, "Litros Agua", "Litros", default=0)
        try:
            litros_total = float(litros_raw) if litros_raw is not None else 0.0
        except (ValueError, TypeError):
            litros_total = 0.0

        dosis_raw = get_val(row, "Dosis gr ó cc x Litro", "Dosis gr  cc x Litro", "Dosis", default=0)
        try:
            dosis = float(dosis_raw) if dosis_raw is not None else 0.0
        except (ValueError, TypeError):
            dosis = 0.0

        total_raw = get_val(row, "Total gramo ó cc", "Total gramo  cc", "Total", default=None)
        try:
            total_producto = float(total_raw) if total_raw is not None else (litros_total * dosis)
        except (ValueError, TypeError):
            total_producto = litros_total * dosis

        norm_pname = normalize_text(prod_name)
        base_info = base_catalog.get(norm_pname)
        
        if not base_info:
            for b_key, b_val in base_catalog.items():
                if b_key in norm_pname or norm_pname in b_key:
                    base_info = b_val
                    break

        if not base_info:
            base_info = {
                "codigo": "",
                "nombre": prod_name,
                "um": "KILO" if "SULFATO" in norm_pname or "NITRATO" in norm_pname or "ACIDO" not in norm_pname else "LITRO",
                "frase_h": "Información SGA no registrada en Base.xlsx para este producto.",
                "frase_p": "Consulte la Ficha de Datos de Seguridad y use equipo de protección personal adecuado.",
                "palabra_advertencia": "ATENCIÓN",
                "pictogramas": [
                    {"has_image": False, "url": None, "label": "SIN IMAGEN", "code": None},
                    {"has_image": False, "url": None, "label": "SIN IMAGEN", "code": None},
                    {"has_image": False, "url": None, "label": "SIN IMAGEN", "code": None},
                    {"has_image": False, "url": None, "label": "SIN IMAGEN", "code": None},
                ],
                "tiene_sga": "NO",
                "tiene_hds": "NO",
            }

        sector_bloque = ""
        if cultivo and bloque:
            sector_bloque = f"{cultivo} - Bloque {bloque}"
        elif cultivo:
            sector_bloque = cultivo
        elif bloque:
            sector_bloque = f"Bloque {bloque}"
        else:
            sector_bloque = "General"

        item_counter += 1
        app_item = {
            "id": f"app-{item_counter}",
            "row_num": row_num,
            "fecha": date_info,
            "semana": semana,
            "cultivo": cultivo,
            "bloque": bloque,
            "sector_bloque": sector_bloque,
            "camas": camas,
            "reentrada": reentrada,
            "categoria": categoria,
            "producto": prod_name,
            "litros_total": litros_total,
            "dosis": dosis,
            "total_producto": total_producto,
            "observaciones": observaciones,
            "base_info": base_info,
        }

        app_item["etiquetas"] = generate_tank_labels(app_item)
        applications.append(app_item)

    return applications


def generate_tank_labels(app_item: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Genera las etiquetas desglosadas por tanques de 1,000 Litros y colitas."""
    volumen_total = app_item["litros_total"]
    dosis = app_item["dosis"]
    base_info = app_item["base_info"]
    um = base_info["um"]

    labels = []

    if volumen_total <= 0:
        return labels

    if volumen_total < 1000:
        cant_dosificar = volumen_total * dosis
        label = {
            "id": f"{app_item['id']}-T1",
            "tanque_num": 1,
            "total_tanques": 1,
            "tipo_tanque": "Tanque Único",
            "es_colita": False,
            "litros_tanque": volumen_total,
            "dosis": dosis,
            "cantidad_dosificar": cant_dosificar,
            "unidad": um,
            "producto": app_item["producto"],
            "sector_bloque": app_item["sector_bloque"],
            "reentrada": app_item["reentrada"],
            "categoria": app_item["categoria"],
            "fecha": app_item["fecha"],
            "base_info": base_info,
            "litros_total_lote": volumen_total,
        }
        labels.append(label)
        return labels

    n_completos = int(volumen_total // 1000)
    residuo = volumen_total % 1000
    total_etiquetas = n_completos + (1 if residuo > 0 else 0)

    for i in range(1, n_completos + 1):
        cant_dosificar = 1000.0 * dosis
        label = {
            "id": f"{app_item['id']}-T{i}",
            "tanque_num": i,
            "total_tanques": total_etiquetas,
            "tipo_tanque": f"Tanque {i} de {total_etiquetas} (1,000 L)",
            "es_colita": False,
            "litros_tanque": 1000.0,
            "dosis": dosis,
            "cantidad_dosificar": cant_dosificar,
            "unidad": um,
            "producto": app_item["producto"],
            "sector_bloque": app_item["sector_bloque"],
            "reentrada": app_item["reentrada"],
            "categoria": app_item["categoria"],
            "fecha": app_item["fecha"],
            "base_info": base_info,
            "litros_total_lote": volumen_total,
        }
        labels.append(label)

    if residuo > 0:
        cant_dosificar = residuo * dosis
        label = {
            "id": f"{app_item['id']}-T{total_etiquetas}",
            "tanque_num": total_etiquetas,
            "total_tanques": total_etiquetas,
            "tipo_tanque": f"Colita {total_etiquetas} de {total_etiquetas} ({residuo:.1f} L)".replace(".0 L", " L"),
            "es_colita": True,
            "litros_tanque": residuo,
            "dosis": dosis,
            "cantidad_dosificar": cant_dosificar,
            "unidad": um,
            "producto": app_item["producto"],
            "sector_bloque": app_item["sector_bloque"],
            "reentrada": app_item["reentrada"],
            "categoria": app_item["categoria"],
            "fecha": app_item["fecha"],
            "base_info": base_info,
            "litros_total_lote": volumen_total,
        }
        labels.append(label)

    return labels


class DataManager:
    """Administrador en memoria de los datos cargados con soporte de recarga en caliente."""
    def __init__(self, data_dir: Optional[str] = None):
        self.data_dir = find_data_directory(data_dir)
        self.base_catalog: Dict[str, Dict[str, Any]] = {}
        self.applications: List[Dict[str, Any]] = []
        self.last_loaded: Optional[datetime.datetime] = None
        self.available_dates: List[Dict[str, str]] = []
        self.available_products: List[str] = []
        self.load_all()

    def load_all(self):
        picto_dir = self.data_dir / "Picto"
        if not picto_dir.exists():
            picto_dir = Path(__file__).resolve().parent / "Picto"

        base_path = self.data_dir / "Base.xlsx"
        app_path = self.data_dir / "aplicacion.xlsm"

        if not base_path.exists():
            raise FileNotFoundError(f"No se encontró Base.xlsx en {self.data_dir}")
        if not app_path.exists():
            raise FileNotFoundError(f"No se encontró aplicacion.xlsm en {self.data_dir}")

        self.base_catalog = load_base_catalog(base_path, picto_dir)
        self.applications = load_applications(app_path, self.base_catalog)
        self.last_loaded = datetime.datetime.now()

        dates_seen = set()
        dates_list = []
        products_seen = set()

        for item in self.applications:
            iso_d = item["fecha"]["iso"]
            if iso_d and iso_d not in dates_seen:
                dates_seen.add(iso_d)
                dates_list.append(item["fecha"])
            
            p_name = item["producto"]
            if p_name and p_name not in products_seen:
                products_seen.add(p_name)

        self.available_dates = sorted(dates_list, key=lambda x: x["iso"], reverse=True)
        self.available_products = sorted(list(products_seen))

    def get_summary(self) -> Dict[str, Any]:
        total_apps = len(self.applications)
        total_labels = sum(len(a["etiquetas"]) for a in self.applications)
        return {
            "data_directory": str(self.data_dir),
            "last_loaded": self.last_loaded.strftime("%Y-%m-%d %H:%M:%S") if self.last_loaded else None,
            "total_applications": total_apps,
            "total_labels": total_labels,
            "total_products_in_catalog": len(self.base_catalog),
            "available_dates": self.available_dates,
            "available_products": self.available_products,
        }
